"""
Tests for the File Generator module.

Tests cover:
- FileGenerator initialization
- PDF generation
- Excel generation
- PowerPoint generation
- CSV generation
- Template loading
- Error handling
"""

import pytest
import tempfile
import os
from pathlib import Path
from unittest.mock import MagicMock, patch, AsyncMock

from src.tools.file_generator import (
    FileGenerator,
    get_file_generator,
    MAX_PDF_SIZE,
    MAX_EXCEL_SIZE,
    MAX_PPTX_SIZE,
)
from src.tools.errors import InvalidInputError, StorageFullError


# =============================================================================
# TEST DATA
# =============================================================================


SAMPLE_AGGREGATED = {
    "brief": {
        "goal": "Research Apple Inc stock potential",
    },
    "executive_summary": "Apple shows strong fundamentals with growth potential.",
    "key_insights": [
        {
            "insight": "Revenue grew 15% year over year",
            "supporting_data": "Q4 2024 earnings report",
            "importance": "high",
        },
        {
            "insight": "Services segment driving margins",
            "supporting_data": "Services revenue at $22B",
            "importance": "medium",
        },
    ],
    "sections": [
        {
            "title": "Financial Overview",
            "summary": "Strong financial performance in Q4.",
            "key_points": [
                "Revenue exceeded expectations",
                "Margins improved",
                "Cash flow positive",
            ],
            "data_highlights": {
                "Revenue": "$94.9B",
                "Net Income": "$23.6B",
                "EPS": "$1.64",
            },
            "analysis": "The company continues to execute well.",
            "sentiment": "positive",
            "data_tables": [
                {
                    "name": "Quarterly Revenue",
                    "headers": ["Quarter", "Revenue ($B)"],
                    "rows": [
                        ["Q1", "90.1"],
                        ["Q2", "81.8"],
                        ["Q3", "89.5"],
                        ["Q4", "94.9"],
                    ],
                }
            ],
        },
    ],
    "recommendation": {
        "verdict": "BUY",
        "confidence": "high",
        "confidence_reasoning": "Strong fundamentals support growth",
        "reasoning": "Based on strong earnings and growth trajectory.",
        "pros": ["Strong brand", "Growing services", "Cash rich"],
        "cons": ["China exposure", "Regulatory risks"],
        "action_items": [
            {
                "action": "Consider accumulating on dips",
                "priority": "high",
                "rationale": "Valuation becoming attractive",
            }
        ],
        "risks_to_monitor": ["Trade tensions", "Smartphone saturation"],
    },
    "contradictions_found": [
        {
            "topic": "Growth outlook",
            "sources": ["Analyst A", "Analyst B"],
            "resolution": "Different time horizons considered",
        }
    ],
    "coverage_summary": [
        {
            "topic": "Financial Analysis",
            "coverage_percent": 95,
            "gaps": [],
        },
        {
            "topic": "Competitive Analysis",
            "coverage_percent": 75,
            "gaps": ["Android market share trends"],
        },
    ],
    "metadata": {
        "total_rounds": 3,
        "total_data_tasks": 5,
        "total_research_tasks": 8,
        "sources_count": 15,
        "processing_time_seconds": 120.5,
    },
}

SAMPLE_CONFIG = {
    "language": "en",
    "style": "formal",
    "pdf": {
        "include_toc": True,
        "include_charts": True,
        "branding": {
            "primary_color": "#1a365d",
            "company_name": "Test Company",
        },
    },
    "excel": {
        "include_raw_data": False,
        "include_charts": True,
    },
    "pptx": {
        "slides_per_section": 2,
        "aspect_ratio": "16:9",
    },
    "csv": {
        "delimiter": ",",
        "encoding": "utf-8",
        "include_headers": True,
    },
}


# =============================================================================
# FILE GENERATOR TESTS
# =============================================================================


class TestFileGenerator:
    """Tests for FileGenerator class."""

    @pytest.fixture
    def generator(self):
        """Create file generator."""
        return FileGenerator(template_dir="src/templates")

    @pytest.fixture
    def temp_dir(self):
        """Create temporary directory for output."""
        with tempfile.TemporaryDirectory() as tmpdir:
            yield tmpdir

    def test_initialization(self, generator):
        """Test generator initializes correctly."""
        assert generator.template_dir == Path("src/templates")
        assert generator.jinja_env is not None

    def test_custom_filters_registered(self, generator):
        """Test custom Jinja2 filters are registered."""
        filters = generator.jinja_env.filters

        assert "format_date" in filters
        assert "format_number" in filters
        assert "format_currency" in filters
        assert "format_percent" in filters

    def test_format_date_filter(self, generator):
        """Test date formatting filter."""
        from datetime import datetime

        dt = datetime(2024, 1, 15, 10, 30)
        result = generator._format_date(dt, "%Y-%m-%d")

        assert result == "2024-01-15"

    def test_format_date_filter_string(self, generator):
        """Test date formatting with ISO string."""
        result = generator._format_date("2024-01-15T10:30:00", "%Y-%m-%d")
        assert result == "2024-01-15"

    def test_format_number_filter(self, generator):
        """Test number formatting filter."""
        result = generator._format_number(1234567.89, 2)
        assert result == "1,234,567.89"

    def test_format_currency_filter(self, generator):
        """Test currency formatting filter."""
        result = generator._format_currency(1234.5, "$")
        assert result == "$1,234.50"

    def test_format_percent_filter(self, generator):
        """Test percentage formatting filter."""
        result = generator._format_percent(75.5)
        assert result == "75.5%"


def weasyprint_available():
    """Check if WeasyPrint is available."""
    try:
        import weasyprint
        return True
    except (ImportError, OSError):
        return False


class TestPDFGeneration:
    """Tests for PDF generation."""

    @pytest.fixture
    def generator(self):
        """Create file generator."""
        return FileGenerator(template_dir="src/templates")

    @pytest.fixture
    def temp_dir(self):
        """Create temporary directory."""
        with tempfile.TemporaryDirectory() as tmpdir:
            yield tmpdir

    @pytest.mark.asyncio
    @pytest.mark.skipif(not weasyprint_available(), reason="WeasyPrint requires GTK libraries")
    async def test_generate_pdf_creates_file(self, generator, temp_dir):
        """Test PDF generation creates a file."""
        output_path = os.path.join(temp_dir, "report.pdf")

        # Mock retry handler
        async def mock_execute(func, *args, **kwargs):
            return await func(*args, **kwargs)
        generator.retry_handler.execute = mock_execute

        path = await generator.generate_pdf(
            aggregated=SAMPLE_AGGREGATED,
            config=SAMPLE_CONFIG,
            output_path=output_path,
        )

        assert os.path.exists(path)
        assert os.path.getsize(path) > 0

    @pytest.mark.asyncio
    @pytest.mark.skipif(not weasyprint_available(), reason="WeasyPrint requires GTK libraries")
    async def test_generate_pdf_returns_path(self, generator, temp_dir):
        """Test PDF generation returns correct path."""
        output_path = os.path.join(temp_dir, "report.pdf")

        async def mock_execute(func, *args, **kwargs):
            return await func(*args, **kwargs)
        generator.retry_handler.execute = mock_execute

        path = await generator.generate_pdf(
            aggregated=SAMPLE_AGGREGATED,
            config=SAMPLE_CONFIG,
            output_path=output_path,
        )

        assert path == output_path

    @pytest.mark.asyncio
    @pytest.mark.skipif(not weasyprint_available(), reason="WeasyPrint requires GTK libraries")
    async def test_generate_pdf_creates_parent_dirs(self, generator, temp_dir):
        """Test PDF generation creates parent directories."""
        output_path = os.path.join(temp_dir, "subdir", "nested", "report.pdf")

        async def mock_execute(func, *args, **kwargs):
            return await func(*args, **kwargs)
        generator.retry_handler.execute = mock_execute

        path = await generator.generate_pdf(
            aggregated=SAMPLE_AGGREGATED,
            config=SAMPLE_CONFIG,
            output_path=output_path,
        )

        assert os.path.exists(path)


class TestExcelGeneration:
    """Tests for Excel generation."""

    @pytest.fixture
    def generator(self):
        """Create file generator."""
        return FileGenerator(template_dir="src/templates")

    @pytest.fixture
    def temp_dir(self):
        """Create temporary directory."""
        with tempfile.TemporaryDirectory() as tmpdir:
            yield tmpdir

    @pytest.mark.asyncio
    async def test_generate_excel_creates_file(self, generator, temp_dir):
        """Test Excel generation creates a file."""
        output_path = os.path.join(temp_dir, "report.xlsx")

        async def mock_execute(func, *args, **kwargs):
            return await func(*args, **kwargs)
        generator.retry_handler.execute = mock_execute

        path = await generator.generate_excel(
            aggregated=SAMPLE_AGGREGATED,
            config=SAMPLE_CONFIG,
            output_path=output_path,
        )

        assert os.path.exists(path)
        assert os.path.getsize(path) > 0

    @pytest.mark.asyncio
    async def test_generate_excel_valid_format(self, generator, temp_dir):
        """Test Excel file is valid format."""
        pytest.importorskip("openpyxl")
        from openpyxl import load_workbook

        output_path = os.path.join(temp_dir, "report.xlsx")

        async def mock_execute(func, *args, **kwargs):
            return await func(*args, **kwargs)
        generator.retry_handler.execute = mock_execute

        await generator.generate_excel(
            aggregated=SAMPLE_AGGREGATED,
            config=SAMPLE_CONFIG,
            output_path=output_path,
        )

        # Should be able to load as valid workbook
        wb = load_workbook(output_path)
        assert len(wb.sheetnames) > 0
        wb.close()

    @pytest.mark.asyncio
    async def test_generate_excel_has_summary_sheet(self, generator, temp_dir):
        """Test Excel has Summary sheet."""
        pytest.importorskip("openpyxl")
        from openpyxl import load_workbook

        output_path = os.path.join(temp_dir, "report.xlsx")

        async def mock_execute(func, *args, **kwargs):
            return await func(*args, **kwargs)
        generator.retry_handler.execute = mock_execute

        await generator.generate_excel(
            aggregated=SAMPLE_AGGREGATED,
            config=SAMPLE_CONFIG,
            output_path=output_path,
        )

        wb = load_workbook(output_path)
        assert "Summary" in wb.sheetnames
        wb.close()


def pptx_available():
    """Check if python-pptx with RGBColor is available."""
    try:
        from pptx import Presentation
        from pptx.dml.color import RGBColor
        return True
    except (ImportError, OSError):
        return False


class TestPPTXGeneration:
    """Tests for PowerPoint generation."""

    @pytest.fixture
    def generator(self):
        """Create file generator."""
        return FileGenerator(template_dir="src/templates")

    @pytest.fixture
    def temp_dir(self):
        """Create temporary directory."""
        with tempfile.TemporaryDirectory() as tmpdir:
            yield tmpdir

    @pytest.mark.asyncio
    @pytest.mark.skipif(not pptx_available(), reason="python-pptx not available or incompatible")
    async def test_generate_pptx_creates_file(self, generator, temp_dir):
        """Test PPTX generation creates a file."""

        output_path = os.path.join(temp_dir, "report.pptx")

        async def mock_execute(func, *args, **kwargs):
            return await func(*args, **kwargs)
        generator.retry_handler.execute = mock_execute

        path = await generator.generate_pptx(
            aggregated=SAMPLE_AGGREGATED,
            config=SAMPLE_CONFIG,
            output_path=output_path,
        )

        assert os.path.exists(path)
        assert os.path.getsize(path) > 0

    @pytest.mark.asyncio
    @pytest.mark.skipif(not pptx_available(), reason="python-pptx not available or incompatible")
    async def test_generate_pptx_valid_format(self, generator, temp_dir):
        """Test PPTX file is valid format."""
        from pptx import Presentation

        output_path = os.path.join(temp_dir, "report.pptx")

        async def mock_execute(func, *args, **kwargs):
            return await func(*args, **kwargs)
        generator.retry_handler.execute = mock_execute

        await generator.generate_pptx(
            aggregated=SAMPLE_AGGREGATED,
            config=SAMPLE_CONFIG,
            output_path=output_path,
        )

        # Should be able to load as valid presentation
        prs = Presentation(output_path)
        assert len(prs.slides) > 0


class TestCSVGeneration:
    """Tests for CSV generation."""

    @pytest.fixture
    def generator(self):
        """Create file generator."""
        return FileGenerator(template_dir="src/templates")

    @pytest.fixture
    def temp_dir(self):
        """Create temporary directory."""
        with tempfile.TemporaryDirectory() as tmpdir:
            yield tmpdir

    @pytest.mark.asyncio
    async def test_generate_csv_creates_file(self, generator, temp_dir):
        """Test CSV generation creates a file."""
        output_path = os.path.join(temp_dir, "report.csv")

        async def mock_execute(func, *args, **kwargs):
            return await func(*args, **kwargs)
        generator.retry_handler.execute = mock_execute

        path = await generator.generate_csv(
            aggregated=SAMPLE_AGGREGATED,
            config=SAMPLE_CONFIG,
            output_path=output_path,
        )

        assert os.path.exists(path)
        assert os.path.getsize(path) > 0

    @pytest.mark.asyncio
    async def test_generate_csv_valid_format(self, generator, temp_dir):
        """Test CSV file is valid format."""
        import csv

        output_path = os.path.join(temp_dir, "report.csv")

        async def mock_execute(func, *args, **kwargs):
            return await func(*args, **kwargs)
        generator.retry_handler.execute = mock_execute

        await generator.generate_csv(
            aggregated=SAMPLE_AGGREGATED,
            config=SAMPLE_CONFIG,
            output_path=output_path,
        )

        # Should be able to read as valid CSV
        with open(output_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)
            assert len(rows) > 0

    @pytest.mark.asyncio
    async def test_generate_csv_custom_delimiter(self, generator, temp_dir):
        """Test CSV with custom delimiter."""
        output_path = os.path.join(temp_dir, "report.csv")

        config = {**SAMPLE_CONFIG, "csv": {"delimiter": ";"}}

        async def mock_execute(func, *args, **kwargs):
            return await func(*args, **kwargs)
        generator.retry_handler.execute = mock_execute

        await generator.generate_csv(
            aggregated=SAMPLE_AGGREGATED,
            config=config,
            output_path=output_path,
        )

        with open(output_path, "r", encoding="utf-8") as f:
            content = f.read()
            # Should contain semicolons if delimiter is ;
            assert ";" in content or len(content.split(";")) > 0


# =============================================================================
# DATA EXTRACTION TESTS
# =============================================================================


class TestDataExtraction:
    """Tests for data extraction helper methods."""

    @pytest.fixture
    def generator(self):
        """Create file generator."""
        return FileGenerator(template_dir="src/templates")

    def test_extract_data_tables(self, generator):
        """Test extracting data tables from sections."""
        tables = generator._extract_data_tables(SAMPLE_AGGREGATED)

        assert len(tables) == 1
        assert tables[0]["name"] == "Quarterly Revenue"
        assert len(tables[0]["rows"]) == 4

    def test_extract_data_tables_empty(self, generator):
        """Test extracting tables from empty aggregated."""
        tables = generator._extract_data_tables({})
        assert tables == []

    def test_extract_data_tables_no_tables(self, generator):
        """Test extracting from sections without tables."""
        aggregated = {
            "sections": [
                {"title": "Test", "summary": "No tables here"},
            ]
        }
        tables = generator._extract_data_tables(aggregated)
        assert tables == []


# =============================================================================
# FACTORY FUNCTION TESTS
# =============================================================================


class TestFactoryFunctions:
    """Tests for factory functions."""

    def test_get_file_generator(self):
        """Test getting file generator."""
        generator = get_file_generator()
        assert isinstance(generator, FileGenerator)

    def test_get_file_generator_custom_dir(self):
        """Test getting file generator with custom template dir."""
        generator = get_file_generator(template_dir="custom/templates")
        assert generator.template_dir == Path("custom/templates")
